from openpyxl import Workbook
import smtplib
import email.message
from dotenv import load_dotenv
import os

def cria_excel_tipo(tipo, data_inicio, data_fim, avaliacoes):
    workbook = Workbook()

    sheet = workbook.active
    sheet.title = "Relatorio"

    headers = ['data', 'turno', 'satisfação', 'comentário']

    for col_numero, header in enumerate(headers, start=1):
        sheet.cell(row=1, column=col_numero, value=header)
    
    quantidade_avaliacoes = len(avaliacoes)

    for nmr, linha_numero in enumerate(range(2, quantidade_avaliacoes + 2), start=2):
        data = avaliacoes[nmr - 2][0]
        satisfacao = get_emogi(avaliacoes[nmr - 2][1])
        comentarios = avaliacoes[nmr - 2][2]
        turno = avaliacoes[nmr - 2][3]

        sheet.cell(row=linha_numero, column=1, value=data)
        sheet.cell(row=linha_numero, column=2, value=turno)        
        sheet.cell(row=linha_numero, column=3, value=satisfacao)
        sheet.cell(row=linha_numero, column=4, value=comentarios)

    if data_inicio and data_fim:
        nome_arquivo = f"Relatorio {tipo} {data_inicio} - {data_fim}.xlsx"
    elif data_inicio:
        nome_arquivo = f"Relatorio {tipo} {data_inicio}.xlsx"
    elif data_fim:
        nome_arquivo = f"Relatorio {tipo} {data_fim}.xlsx"
    else:
        nome_arquivo = f"Relatorio {tipo}.xlsx"

    workbook.save(nome_arquivo)
    envio_email = enviar_email('rafaperovano@gmail.com', nome_arquivo)
    
    return envio_email

def get_emogi(satisfacao):
    emojis = { '0': '😡', '1': '😟', '2': '😐', '3': '🙂', '4': '😀' }
    return emojis[satisfacao]

def enviar_email(destinatario, nome_arquivo):  
    load_dotenv()
    senha = os.getenv('SENHA_EMAIL')

    if not senha:
        raise ValueError("A variável de ambiente SENHA_EMAIL não foi encontrada. Certifique-se de que está configurada corretamente.")

    msg = email.message.EmailMessage()
    msg['Subject'] = f"{nome_arquivo} das avaliações"
    msg['From'] = "rafaeltestepython@gmail.com"
    msg['To'] = destinatario
    msg.set_content("Segue o arquivo em anexo.")

    if not os.path.exists(nome_arquivo):
        print("Arquivo não encontrado para anexo.")
        return

    with open(nome_arquivo, "rb") as f:
        msg.add_attachment(f.read(), maintype="application",
        subtype="vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        filename=nome_arquivo)

    try:
        with smtplib.SMTP("smtp.gmail.com", 587) as server:
            server.starttls()
            server.login(msg['From'], senha)
            server.send_message(msg)

        os.remove(nome_arquivo)

        return True
    
    except Exception as e:
        return False